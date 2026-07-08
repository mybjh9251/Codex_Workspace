from __future__ import annotations

import csv
import logging
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook

from models import AppConfig, SatelliteRequest

SUPPORTED_INPUT_SUFFIXES = {".xlsx", ".csv"}


def read_satellite_requests(
    config: AppConfig,
    logger: logging.Logger,
) -> list[SatelliteRequest]:
    input_path = config.input_satellite_path
    if not input_path.exists():
        raise FileNotFoundError(f"Input file was not found: {input_path}")

    suffix = input_path.suffix.lower()

    if suffix == ".xlsx":
        return _read_xlsx_satellite_requests(config, logger, input_path)

    if suffix == ".csv":
        return _read_csv_satellite_requests(config, logger, input_path)

    supported = ", ".join(sorted(SUPPORTED_INPUT_SUFFIXES))
    raise ValueError(
        f"Unsupported input file format: {input_path.name}. "
        f"Supported formats: {supported}."
    )


def _read_xlsx_satellite_requests(
    config: AppConfig,
    logger: logging.Logger,
    input_path: Path,
) -> list[SatelliteRequest]:
    workbook = load_workbook(
        filename=input_path,
        read_only=True,
        data_only=True,
    )
    try:
        if config.sheet_name not in workbook.sheetnames:
            raise ValueError(
                f"Sheet '{config.sheet_name}' was not found in {input_path.name}."
            )

        worksheet = workbook[config.sheet_name]
        header_row = next(
            worksheet.iter_rows(min_row=1, max_row=1, values_only=True),
            None,
        )
        if header_row is None:
            raise ValueError("The worksheet does not contain a header row.")

        sat_name_index, norad_id_index = _resolve_header_indexes(config, header_row)

        requests: list[SatelliteRequest] = []
        for row_number, row_values in enumerate(
            worksheet.iter_rows(min_row=2, values_only=True),
            start=2,
        ):
            _append_satellite_request(
                requests,
                logger,
                row_number,
                row_values,
                sat_name_index,
                norad_id_index,
            )

        return requests
    finally:
        workbook.close()


def _read_csv_satellite_requests(
    config: AppConfig,
    logger: logging.Logger,
    input_path: Path,
) -> list[SatelliteRequest]:
    with input_path.open("r", encoding="utf-8-sig", newline="") as csv_file:
        sample = csv_file.read(4096)
        csv_file.seek(0)
        dialect = _detect_csv_dialect(sample)
        reader = csv.reader(csv_file, dialect)

        header_row = next(reader, None)
        if header_row is None:
            raise ValueError("The CSV file does not contain a header row.")

        sat_name_index, norad_id_index = _resolve_header_indexes(config, header_row)

        requests: list[SatelliteRequest] = []
        for row_number, row_values in enumerate(reader, start=2):
            _append_satellite_request(
                requests,
                logger,
                row_number,
                row_values,
                sat_name_index,
                norad_id_index,
            )

    return requests


def _detect_csv_dialect(sample: str) -> csv.Dialect:
    if not sample:
        return csv.get_dialect("excel")

    try:
        return csv.Sniffer().sniff(sample, delimiters=",\t;")
    except csv.Error:
        return csv.get_dialect("excel")


def _resolve_header_indexes(
    config: AppConfig,
    header_row: Iterable[object],
) -> tuple[int, int]:
    header_map = {
        _normalize_header_name(value): index
        for index, value in enumerate(header_row)
    }
    required_headers = [config.sat_name_header, config.norad_id_header]
    missing_headers = [
        header
        for header in required_headers
        if _normalize_header_name(header) not in header_map
    ]
    if missing_headers:
        raise ValueError(
            "Missing required headers: " + ", ".join(missing_headers)
        )

    sat_name_index = header_map[_normalize_header_name(config.sat_name_header)]
    norad_id_index = header_map[_normalize_header_name(config.norad_id_header)]
    return sat_name_index, norad_id_index


def _append_satellite_request(
    requests: list[SatelliteRequest],
    logger: logging.Logger,
    row_number: int,
    row_values: Iterable[object],
    sat_name_index: int,
    norad_id_index: int,
) -> None:
    sat_name = _clean_text(_safe_get(row_values, sat_name_index))
    norad_id = _normalize_norad_id(_safe_get(row_values, norad_id_index))

    if not sat_name and not norad_id:
        return

    if not sat_name or not norad_id:
        logger.warning(
            "Skipping row %s because SAT_Name or NORAD ID is empty.",
            row_number,
        )
        return

    requests.append(
        SatelliteRequest(
            row_number=row_number,
            sat_name=sat_name,
            norad_id=norad_id,
        )
    )


def _safe_get(values: Iterable[object], index: int) -> object:
    values_list = list(values)
    if index >= len(values_list):
        return ""
    return values_list[index]


def _normalize_header_name(value: object) -> str:
    return _clean_text(value)


def _clean_text(value: object) -> str:
    if value is None:
        return ""
    return str(value).replace("\u00a0", " ").strip()


def _normalize_norad_id(value: object) -> str:
    if value is None:
        return ""

    if isinstance(value, int):
        return str(value)

    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return ""

    text = _clean_text(value)
    normalized = text.replace(",", "")
    if normalized.isdigit():
        return str(int(normalized))
    return ""
